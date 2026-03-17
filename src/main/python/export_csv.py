# -*- coding: utf-8 -*-
# 最先抑制 trio 的 RuntimeWarning（openai 依赖触发），避免干扰输出
import warnings
warnings.filterwarnings("ignore", category=RuntimeWarning, module=r"trio.*")

"""
根据需求描述调用 AI 生成功能测试用例，并导出为 CSV。
支持多种生成方式（按优先级）：
  1. --json / --json-file：使用预生成的 JSON（如 Cursor IDE 内 Agent 生成结果）
  2. --use-cursor-agent：调用 Cursor Cloud Agents API（需 CURSOR_API_KEY、CURSOR_REPO_URL）
  3. 默认：豆包 / OpenAI 兼容接口（OpenAI 官方、国内代理等）

豆包配置（火山方舟）：
  DOUBAO_API_KEY=你的豆包API_Key
  DOUBAO_MODEL=推理接入点ID（ep-开头）或模型名（如 doubao-1.5-pro-32k）

OpenAI 兼容配置（OpenAI 官方、OneAPI 代理等）：
  OPENAI_API_KEY=你的API_Key
  OPENAI_BASE_URL=可选，用于代理（如 https://api.openai.com/v1 或代理地址）
  OPENAI_MODEL=模型名称，默认 gpt-4o-mini

Cursor Cloud Agent 配置（--use-cursor-agent 时使用）：
  CURSOR_API_KEY=从 Cursor Dashboard 获取
  CURSOR_REPO_URL=GitHub 仓库地址，如 https://github.com/your-org/your-repo
"""
import argparse
import csv
import json
import os
import re
import sys
import time

HEADERS = ["用例编号", "模块", "测试点", "前置条件", "测试步骤", "预期结果", "优先级", "测试结果", "备注"]

DOUBAO_BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
CURSOR_API_URL = "https://api.cursor.com/v0"

# 标题行样式：背景色 RGB(68,114,196)
HEADER_BG_RGB = "4472C4"


def _format_steps_with_newlines(text: str) -> str:
    """将「前置条件」「测试步骤」中的步骤（1. 2、 第2步 等）之间用换行符分隔（等效于单元格内 ALT+Enter）。"""
    if not text or not isinstance(text, str):
        return str(text or "")
    text = text.strip()
    # 先按分号分割（步骤可能用 ; 或 ； 分隔）
    for sep in ["；", ";"]:
        text = text.replace(sep, "\n")
    # 中文逗号后紧跟步骤编号时换行（如 "1. 打开，2. 输入"），避免误伤步骤内逗号
    text = re.sub(r"，\s*(?=(?:\d+[.、．:：)）]|第\d+步))", "\n", text)
    # 在步骤编号前的空白处插入换行：1. 2、 3） 第2步 或 1 打开 2 输入
    text = re.sub(r"\s+(?=(?:\d+[.、．:：)）]|第\d+步)\s*)", "\n", text)
    text = re.sub(r"\s+(?=\d+\s+[^\d\s\n])", "\n", text)  # 如 "1 打开 2 输入"
    # 无空格时的紧凑格式（如「第1步xxx第2步」「1.aaa2.bbb」）：在步骤标记前插入换行
    text = re.sub(r"(?<=[^\n])(?=(?:\d+[.、．:：)）]|第\d+步))", "\n", text)
    # 合并连续的换行
    return re.sub(r"\n{2,}", "\n", text).strip()


def _normalize_step_field(val) -> str:
    """将前置条件/测试步骤统一为字符串，支持列表格式。"""
    if val is None:
        return ""
    if isinstance(val, list):
        return "\n".join(str(x).strip() for x in val if x)
    return str(val).strip()


def _parse_cases_to_rows(cases: list) -> list:
    """将用例 dict 列表转为 CSV 行格式。"""
    rows = [HEADERS]
    for i, c in enumerate(cases, 1):
        if not isinstance(c, dict):
            continue
        pre_raw = _normalize_step_field(c.get("前置条件", "") or "")
        step_raw = _normalize_step_field(c.get("测试步骤", "") or "")
        precondition = _format_steps_with_newlines(pre_raw)
        steps = _format_steps_with_newlines(step_raw)
        row = [
            i,
            c.get("模块", ""),
            c.get("测试点", ""),
            precondition,
            steps,
            c.get("预期结果", ""),
            c.get("优先级", "P2"),
            c.get("测试结果", ""),
            c.get("备注", ""),
        ]
        rows.append(row)
    return rows


def _extract_json_from_text(text: str) -> list:
    """从文本中提取 JSON 数组。"""
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```\s*$", "", text)
    text = text.strip()
    # 尝试匹配 [...]
    match = re.search(r"\[[\s\S]*\]", text)
    if match:
        text = match.group(0)
    return json.loads(text)


def generate_with_cursor_agent(requirement: str) -> list:
    """
    调用 Cursor Cloud Agents API 生成功能测试用例。
    需设置环境变量：CURSOR_API_KEY、CURSOR_REPO_URL。
    """
    api_key = os.environ.get("CURSOR_API_KEY")
    repo_url = os.environ.get("CURSOR_REPO_URL")
    if not api_key or not repo_url:
        raise RuntimeError(
            "使用 Cursor Agent 需设置环境变量：\n"
            "  CURSOR_API_KEY=你的API_Key（从 Cursor Dashboard 获取）\n"
            "  CURSOR_REPO_URL=GitHub 仓库地址，如 https://github.com/your-org/your-repo"
        )

    try:
        import requests
    except ImportError:
        raise RuntimeError("请先安装: pip install requests")

    auth = (api_key, "")
    prompt_text = f"""根据以下需求描述，生成功能测试用例。请只输出一个 JSON 数组，每个用例包含字段：
- 模块
- 测试点
- 前置条件
- 测试步骤（多步用分号；分隔，如：1.xxx；2.xxx；3.xxx）
- 预期结果
- 优先级（P1/P2/P3）

只返回 JSON 数组，不要其他解释文字。示例格式：
[
  {{"模块": "XX", "测试点": "XX", "前置条件": "XX", "测试步骤": "1.xxx；2.xxx；3.xxx", "预期结果": "XX", "优先级": "P1"}},
  ...
]

需求描述：
{requirement}
"""

    # 发起 Agent 任务
    resp = requests.post(
        f"{CURSOR_API_URL}/agents",
        auth=auth,
        json={
            "prompt": {"text": prompt_text},
            "source": {"repository": repo_url},
            "target": {"autoCreatePr": False},
        },
        headers={"Content-Type": "application/json"},
        timeout=30,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"启动 Cursor Agent 失败: {resp.status_code} - {resp.text}")

    data = resp.json()
    agent_id = data.get("id")
    if not agent_id:
        raise RuntimeError("未获取到 Agent ID")

    # 轮询等待完成
    max_wait = 300  # 5 分钟
    step = 5
    for _ in range(0, max_wait, step):
        status_resp = requests.get(f"{CURSOR_API_URL}/agents/{agent_id}", auth=auth, timeout=30)
        if status_resp.status_code != 200:
            raise RuntimeError(f"查询 Agent 状态失败: {status_resp.status_code}")
        status_data = status_resp.json()
        status = status_data.get("status", "")
        if status == "FINISHED":
            break
        if status in ("FAILED", "CANCELLED", "STOPPED"):
            raise RuntimeError(f"Cursor Agent 执行失败: {status}")
        time.sleep(step)
    else:
        raise RuntimeError("Cursor Agent 执行超时")

    # 获取对话内容并解析 JSON
    conv_resp = requests.get(
        f"{CURSOR_API_URL}/agents/{agent_id}/conversation",
        auth=auth,
        timeout=30,
    )
    if conv_resp.status_code != 200:
        raise RuntimeError(f"获取 Agent 对话失败: {conv_resp.status_code}")

    conv = conv_resp.json()
    messages = conv.get("messages", [])
    all_text = ""
    for m in messages:
        if m.get("type") == "assistant_message":
            all_text += m.get("text", "") + "\n"

    try:
        cases = _extract_json_from_text(all_text)
    except (json.JSONDecodeError, re.error) as e:
        raise RuntimeError(f"无法从 Agent 输出中解析 JSON: {e}\n输出片段: {all_text[:500]}")

    if not isinstance(cases, list):
        raise RuntimeError("解析结果应为 JSON 数组")
    return _parse_cases_to_rows(cases)


def generate_test_cases_with_llm(requirement: str) -> list:
    """
    调用 AI 模型生成功能测试用例。
    优先级：豆包（DOUBAO_API_KEY）> OpenAI 兼容接口（OPENAI_API_KEY）。
    注：Cursor 不提供公开 Chat API，请使用 OpenAI 或豆包等。
    """
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", RuntimeWarning)
            from openai import OpenAI
    except ImportError:
        raise RuntimeError("请先安装: pip install openai")

    api_key = os.environ.get("DOUBAO_API_KEY") or os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError(
            "请设置环境变量：\n"
            "  豆包: DOUBAO_API_KEY=你的API_Key，以及 DOUBAO_MODEL=推理接入点ID\n"
            "  OpenAI: OPENAI_API_KEY=你的API_Key（支持 OpenAI 官方或代理）"
        )

    if os.environ.get("DOUBAO_API_KEY"):
        base_url = DOUBAO_BASE_URL
        model = os.environ.get("DOUBAO_MODEL", "")
        if not model:
            raise RuntimeError(
                "使用豆包时需设置 DOUBAO_MODEL（推理接入点ID，如 ep-2024xxxx）\n"
                "请在火山方舟控制台 → 模型推理 → 推理接入点 → 创建接入点 后获取"
            )
    else:
        base_url = os.environ.get("OPENAI_BASE_URL")
        model = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")

    client = OpenAI(api_key=api_key, base_url=base_url or None)

    prompt = f"""请根据以下需求描述，生成功能测试用例。严格按 JSON 数组格式返回，每个用例包含字段：
- 模块
- 测试点
- 前置条件
- 测试步骤（多步用分号；分隔，如：1.xxx；2.xxx；3.xxx）
- 预期结果
- 优先级（P1/P2/P3）

只返回 JSON，不要其他解释文字。示例格式：
[
  {{"模块": "XX", "测试点": "XX", "前置条件": "XX", "测试步骤": "1.xxx；2.xxx；3.xxx", "预期结果": "XX", "优先级": "P1"}},
  ...
]

需求描述：
{requirement}
"""

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", RuntimeWarning)
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
            )
        text = response.choices[0].message.content.strip()
    except Exception as e:
        err_msg = str(e).strip() or repr(e)
        if hasattr(e, "response") and getattr(e, "response", None) is not None:
            try:
                resp = e.response
                body = resp.text if hasattr(resp, "text") else str(resp)
                if body:
                    err_msg += f"\n响应内容: {body[:500]}"
            except Exception:
                pass
        raise RuntimeError(
            f"模型 API 调用失败：{err_msg}\n"
            "提示：请检查 API Key 是否有效、网络是否可达；使用代理时需设置 OPENAI_BASE_URL"
        ) from e

    try:
        cases = _extract_json_from_text(text)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"模型返回内容无法解析为 JSON: {e}\n原始内容片段: {text[:500]}")

    if not isinstance(cases, list):
        raise RuntimeError("模型返回格式错误：应为 JSON 数组")
    return _parse_cases_to_rows(cases)


def load_cases_from_json(json_input: str | None, json_file: str | None) -> list:
    """从 JSON 字符串或文件加载用例并转为 CSV 行格式。"""
    if json_file:
        with open(json_file, "r", encoding="utf-8") as f:
            raw = f.read()
    elif json_input:
        raw = json_input
    else:
        raise ValueError("需提供 --json 或 --json-file")
    cases = _extract_json_from_text(raw)
    if not isinstance(cases, list):
        raise ValueError("JSON 应为数组格式")
    return _parse_cases_to_rows(cases)


def _write_xlsx(rows: list, wb) -> None:
    """将行数据写入 openpyxl 工作簿，并应用格式。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.active
    ws.title = "功能测试用例"

    # 标题行样式：加粗、背景 RGB(68,114,196)、字号+1、白色字体
    header_font = Font(bold=True, size=12, color="FFFFFF")
    body_font = Font(size=11)
    header_fill = PatternFill(start_color=HEADER_BG_RGB, end_color=HEADER_BG_RGB, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 自动列宽
    for c in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # 为含换行的行设置足够行高，确保步骤逐行显示（前置条件列4、测试步骤列5）
    for r in range(2, len(rows) + 1):
        max_lines = 1
        for col in [4, 5]:
            val = rows[r - 1][col - 1] if r - 1 < len(rows) and col <= len(rows[r - 1]) else None
            if val is not None and isinstance(val, str) and "\n" in val:
                max_lines = max(max_lines, val.count("\n") + 1)
        if max_lines > 1:
            ws.row_dimensions[r].height = max(15 * max_lines, 30)


def export_rows_to_xlsx_bytes(rows: list) -> bytes:
    """将行数据导出为 xlsx 字节，供 API 使用。"""
    from openpyxl import Workbook

    wb = Workbook()
    _write_xlsx(rows, wb)
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_csv(rows: list, output_path: str) -> None:
    """
    将行数据导出为文件。
    - 扩展名为 .xlsx 时导出为带格式的 Excel（标题加粗、背景色、框线、步骤换行）
    - 其他扩展名导出为纯 CSV。
    """
    output_path = output_path.strip()
    if output_path.lower().endswith(".xlsx"):
        from openpyxl import Workbook

        wb = Workbook()
        _write_xlsx(rows, wb)
        wb.save(output_path)
    else:
        with open(output_path, "w", encoding="gb18030", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(
        description="根据需求描述生成功能测试用例并导出为 CSV",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python export_csv.py "用户登录功能：支持手机号/邮箱登录，需验证码"
  python export_csv.py --requirement "订单创建流程测试"
  python export_csv.py --use-cursor-agent "支付模块测试"
  python export_csv.py --json-file cases.json -o 测试用例.csv
        """,
    )
    parser.add_argument(
        "requirement",
        nargs="?",
        default=None,
        help="需求描述文本（也可通过 --requirement 传入）",
    )
    parser.add_argument(
        "-r", "--requirement",
        dest="requirement_arg",
        default=None,
        help="需求描述文本",
    )
    parser.add_argument(
        "-o", "--output",
        default="功能测试用例.xlsx",
        help="输出文件路径（默认：功能测试用例.xlsx，支持 .xlsx 带格式或 .csv 纯文本）",
    )
    parser.add_argument(
        "--use-cursor-agent",
        action="store_true",
        help="使用 Cursor Cloud Agents API 生成（需 CURSOR_API_KEY、CURSOR_REPO_URL）",
    )
    parser.add_argument(
        "--json",
        default=None,
        help="直接传入 JSON 格式的用例数组字符串",
    )
    parser.add_argument(
        "--json-file",
        default=None,
        help="从文件读取预生成的 JSON 用例（如 Cursor Agent 生成结果）",
    )
    args = parser.parse_args()

    requirement = (args.requirement_arg or args.requirement or "").strip()
    if args.json or args.json_file:
        try:
            rows = load_cases_from_json(args.json, args.json_file)
        except (ValueError, FileNotFoundError, json.JSONDecodeError) as e:
            print(f"加载 JSON 失败: {e}")
            sys.exit(1)
    elif args.use_cursor_agent:
        if not requirement:
            print("使用 --use-cursor-agent 时需提供需求描述（-r/--requirement 或位置参数）")
            sys.exit(1)
        print("正在调用 Cursor Agent 生成测试用例...")
        try:
            rows = generate_with_cursor_agent(requirement)
        except RuntimeError as e:
            print(f"Cursor Agent 失败: {e}")
            sys.exit(1)
    else:
        if not requirement:
            print("请输入需求描述（多行输入时，输入空行结束）：")
            lines = []
            while True:
                try:
                    line = input()
                except EOFError:
                    break
                if line == "" and lines:
                    break
                if line == "":
                    continue
                lines.append(line)
            requirement = "\n".join(lines).strip()
        if not requirement:
            print("需求描述不能为空")
            sys.exit(1)
        print("正在调用模型生成测试用例...")
        try:
            rows = generate_test_cases_with_llm(requirement)
        except RuntimeError as e:
            print(f"模型调用失败: {e}")
            sys.exit(1)

    export_to_csv(rows, args.output)
    print(f"导出成功：{args.output}（共 {len(rows) - 1} 条用例）")


if __name__ == "__main__":
    main()
